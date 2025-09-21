from flask import Flask, render_template, request, redirect, session, url_for, flash, send_file, jsonify, after_this_request, Response, make_response
chat_history = []
import sqlite3
from datetime import datetime, timedelta
import pandas as pd
from flask_socketio import SocketIO, emit
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import os
import bcrypt
import tempfile
import requests
from io import BytesIO
import re
from requests.auth import HTTPBasicAuth
import json
import hashlib
import uuid
from dotenv import load_dotenv
load_dotenv()  # charge le fichier .env dans les variables d'environnement
from flask_wtf import CSRFProtect
from flask_wtf.csrf import generate_csrf
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address





AIRCALL_API_TOKEN = os.getenv("AIRCALL_API_TOKEN", "")
# ──────────────────────────────────────────────────────────────────────────────
# Helpers: normalisation & MAJ CALL_ID depuis Aircall
# ──────────────────────────────────────────────────────────────────────────────
def _normalize_phone(num: str) -> str:
    return re.sub(r'[^\d+]', '', str(num or '')).strip()

def update_call_id_in_db(phone_number: str) -> str:
    """
    Récupère le dernier call_id depuis Aircall pour un numéro et
    met à jour la colonne CALL_ID de tous les enregistrements clients
    qui ont ce numéro mais un CALL_ID vide.
    Retourne le call_id ('' si rien trouvé).
    """
    clean = _normalize_phone(phone_number)
    call_id = get_last_aircall_id_by_number(clean)  # <- déjà défini dans ton code
    if not call_id:
        return ""

    try:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        # MAJ toutes les lignes avec ce téléphone et CALL_ID vide
        c.execute("""
            UPDATE clients
               SET CALL_ID = ?
             WHERE REPLACE(REPLACE(REPLACE(TELEPHONE,' ',''),'-',''),'.','') LIKE ?
               AND (CALL_ID IS NULL OR CALL_ID = '')
        """, (call_id, f"%{clean}%"))
        conn.commit()
    finally:
        conn.close()
    return call_id

# ──────────────────────────────────────────────────────────────────────────────
# Fonctions Aircall
# ──────────────────────────────────────────────────────────────────────────────
def find_recording_for_phone_number(phone_number):
    """
    Trouve l'URL d'enregistrement la plus récente pour un numéro donné.
    Méthode Aircall conseillée : 1) trouver le contact, 2) lister ses appels.
    """
    print(f"\n🎯 Recherche d'enregistrement pour : {phone_number}")

    # 1) Chercher le contact
    contact_url = "https://api.aircall.io/v1/contacts"
    clean_number = re.sub(r'[^\d+]', '', str(phone_number)).strip()
    params_contact = {"phone_number": clean_number}
    print(f"   1️⃣ Recherche du contact pour '{clean_number}'")

    try:
        response_contact = requests.get(
            contact_url,
            params=params_contact,
            auth=HTTPBasicAuth(API_ID, API_TOKEN),
            timeout=15
        )
        response_contact.raise_for_status()
        contacts_data = response_contact.json()
        if not contacts_data.get('contacts'):
            print("   ❌ Aucun contact trouvé.")
            return None

        contact = contacts_data['contacts'][0]
        contact_id = contact['id']
        print(f"   ✅ Contact trouvé : ID={contact_id} (nom={contact.get('name', 'N/A')})")
    except requests.exceptions.RequestException as e:
        print(f"   💥 Erreur recherche contact : {e}")
        return None

    # 2) Lister les appels récents du contact
    calls_url = "https://api.aircall.io/v1/calls"
    from_date = (datetime.now() - timedelta(days=90)).isoformat()
    params_calls = {
        "contact_id": contact_id,
        "order": "desc",
        "per_page": 30,
        "from": from_date
    }
    print(f"   2️⃣ Liste des appels récents pour contact {contact_id}")

    try:
        response_calls = requests.get(
            calls_url,
            params=params_calls,
            auth=HTTPBasicAuth(API_ID, API_TOKEN),
            timeout=15
        )
        response_calls.raise_for_status()
        calls_data = response_calls.json()
        calls = calls_data.get('calls', [])
        if not calls:
            print("   ❌ Aucun appel récent.")
            return None

        print(f"   ✅ {len(calls)} appels trouvés. Recherche d'un enregistrement...")
        for call in calls:
            if call.get('recording'):
                print(f"   🎵 Enregistrement trouvé pour appel {call.get('id')} (started_at={call.get('started_at')})")
                return call['recording']  # URL temporaire fournie par Aircall
        print("   ❌ Aucun enregistrement disponible parmi les appels.")
        return None
    except requests.exceptions.RequestException as e:
        print(f"   💥 Erreur liste appels : {e}")
        return None


def get_last_aircall_id_by_number(phone_number):
    """
    Récupère l'ID du dernier appel pour un numéro (utile pour CALL_ID).
    Même logique : 1) contact, 2) appels (le plus récent).
    Renvoie une chaîne vide si rien.
    """
    try:
        # 1) Contact
        contact_url = "https://api.aircall.io/v1/contacts"
        clean_number = re.sub(r'[^\d+]', '', str(phone_number)).strip()
        r1 = requests.get(
            contact_url,
            params={"phone_number": clean_number},
            auth=HTTPBasicAuth(API_ID, API_TOKEN),
            timeout=15
        )
        r1.raise_for_status()
        data1 = r1.json()
        if not data1.get("contacts"):
            return ""

        contact_id = data1["contacts"][0]["id"]

        # 2) Derniers appels du contact
        calls_url = "https://api.aircall.io/v1/calls"
        from_date = (datetime.now() - timedelta(days=90)).isoformat()
        r2 = requests.get(
            calls_url,
            params={"contact_id": contact_id, "order": "desc", "per_page": 1, "from": from_date},
            auth=HTTPBasicAuth(API_ID, API_TOKEN),
            timeout=15
        )
        r2.raise_for_status()
        calls = r2.json().get("calls", [])
        if not calls:
            return ""
        return str(calls[0].get("id") or "")
    except Exception as e:
        print(f"[get_last_aircall_id_by_number] Erreur: {e}")
        return ""

# ──────────────────────────────────────────────────────────────────────────────
# Flask
# ──────────────────────────────────────────────────────────────────────────────
app = Flask(__name__)
# Secret Key lue depuis .env (obligatoire en prod). En dev, on tolère un fallback généré.
app.secret_key = os.getenv("SECRET_KEY") or os.urandom(32)
socketio = SocketIO(app)
# --- Rate Limiting ---
# Stockage en mémoire (suffisant pour une instance simple). Pour du multi-instance, utiliser Redis.
limiter = Limiter(
    get_remote_address,              # clé = IP du client
    app=app,
    default_limits=["200 per day", "50 per hour"],  # limites globales "soft"
    storage_uri="memory://",
)


DB_NAME = os.getenv("DB_NAME", "crm_clients.db")
UPLOAD_FOLDER = os.getenv("UPLOAD_FOLDER", os.path.join('static', 'uploads'))
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER, exist_ok=True)
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # 5 Mo

# --- CSRF ---
app.config['WTF_CSRF_ENABLED'] = True            # activé explicitement (par défaut True, mais on force)
app.config['WTF_CSRF_TIME_LIMIT'] = None         # pas d'expiration du token pendant la session (pratique en dev)
csrf = CSRFProtect(app)

# Rendre le token CSRF accessible dans les templates : {{ csrf_token() }}
@app.context_processor
def inject_csrf_token():
    return dict(csrf_token=lambda: generate_csrf())

# ──────────────────────────────────────────────────────────────────────────────
# Base de données
# ──────────────────────────────────────────────────────────────────────────────
def creer_table():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # Table clients
    c.execute("""
        CREATE TABLE IF NOT EXISTS clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            DATE_SIGNATURE TEXT NOT NULL,
            CIVILITE_CLIENT TEXT DEFAULT '',
            NOM_CLIENT TEXT DEFAULT '',
            PRENOM_CLIENT TEXT DEFAULT '',
            TELEPHONE TEXT DEFAULT '',
            STATUT TEXT DEFAULT '',
            AGENT TEXT DEFAULT '',
            DEUXIEME_ADRESSE TEXT DEFAULT '',
            TROISIEME_ADRESSE TEXT DEFAULT '',
            TYPE_OFFRE TEXT DEFAULT '',
            CREE_PAR TEXT,
            MODIFIE_PAR TEXT,
            DATE_MODIF TEXT
        )
    """)

    # Colonnes additionnelles
    champs_a_ajouter = [
        "campagne_id INTEGER DEFAULT 1",
        "TITRE TEXT DEFAULT ''",
        "NOM_VENDEUR TEXT DEFAULT ''",
        "PRENOM_VENDEUR TEXT DEFAULT ''",
        "TELEPHONE_VENDEUR TEXT DEFAULT ''",
        "N_CONTRAT TEXT DEFAULT ''",
        "N_REFERENCE TEXT DEFAULT ''",
        "VALIDATION_PRODUIT1 TEXT DEFAULT ''",
        "STATUT_PRODUIT1 TEXT DEFAULT ''",
        "VALIDATION_PRODUIT2 TEXT DEFAULT ''",
        "STATUT_PRODUIT2 TEXT DEFAULT ''",
        "VALIDATION_PRODUIT3 TEXT DEFAULT ''",
        "STATUT_PRODUIT3 TEXT DEFAULT ''",
        "EXTRANET TEXT DEFAULT ''"
    ]
    for champ in champs_a_ajouter:
        try:
            c.execute(f"ALTER TABLE clients ADD COLUMN {champ}")
        except sqlite3.OperationalError:
            pass

    # Table agents
    c.execute("""
        CREATE TABLE IF NOT EXISTS agents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            NOM TEXT NOT NULL UNIQUE,
            LOGIN TEXT NOT NULL UNIQUE,
            MDP TEXT NOT NULL,
            ROLE TEXT NOT NULL DEFAULT 'agent'
        )
    """)
    try:
        c.execute("ALTER TABLE agents ADD COLUMN campagne_id INTEGER DEFAULT 1")
    except sqlite3.OperationalError:
        pass

    # Journal connexions
    c.execute("""
        CREATE TABLE IF NOT EXISTS journal_connexions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            agent_nom TEXT NOT NULL,
            date_connexion TEXT NOT NULL,
            page TEXT,
            type_event TEXT DEFAULT 'connexion'
        )
    """)

    # Historique clients
    c.execute("""
        CREATE TABLE IF NOT EXISTS historique_clients (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            client_id INTEGER,
            date_modif TEXT,
            agent TEXT,
            champ_modifie TEXT,
            ancienne_valeur TEXT,
            nouvelle_valeur TEXT
        )
    """)

    # Campagnes
    c.execute("""
        CREATE TABLE IF NOT EXISTS campagnes (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nom TEXT NOT NULL,
            type_export TEXT NOT NULL
        )
    """)

    # Valeurs par défaut des campagnes
    c.execute("SELECT COUNT(*) FROM campagnes WHERE nom='EXOSPHERE_SFR'")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO campagnes (nom, type_export) VALUES (?, ?)", ('EXOSPHERE_SFR', 'simple'))
    c.execute("SELECT COUNT(*) FROM campagnes WHERE nom='VALANDRE'")
    if c.fetchone()[0] == 0:
        c.execute("INSERT INTO campagnes (nom, type_export) VALUES (?, ?)", ('VALANDRE', 'special'))

    # Produits Valandre
    produits = ["STRATO", "LSR", "PRESSE", "ENI", "SERENITY", "PROTEC_ALLIANCE", "WEKIWI"]
    for prod in produits:
        try:
            c.execute(f"ALTER TABLE clients ADD COLUMN {prod}_NUM TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        try:
            c.execute(f"ALTER TABLE clients ADD COLUMN {prod}_STATUT TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        try:
            c.execute(f"ALTER TABLE clients ADD COLUMN {prod}_REMARQUE TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass
        try:
            c.execute("ALTER TABLE agents ADD COLUMN photo TEXT")
        except sqlite3.OperationalError:
            pass
        try:
            c.execute("ALTER TABLE clients ADD COLUMN CALL_ID TEXT DEFAULT ''")
        except sqlite3.OperationalError:
            pass

    # Chat
    c.execute("""
        CREATE TABLE IF NOT EXISTS chat_messages (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user TEXT,
            message TEXT,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP
        )
    """)

    conn.commit()
    conn.close()

def get_agents():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT NOM FROM agents")
    agents = [row[0] for row in c.fetchall()]
    conn.close()
    return agents

def get_campagnes():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id, nom FROM campagnes")
    campagnes = c.fetchall()
    conn.close()
    return campagnes

creer_table()

# ──────────────────────────────────────────────────────────────────────────────
# Authentification
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/login', methods=['GET', 'POST'])
@limiter.limit("5 per minute; 20 per hour", methods=["POST"], error_message="Trop de tentatives. Réessayez dans une minute.")
def login():
    if request.method == 'POST':
        if 'agent_nom' in session:
            return redirect(url_for('index'))
        login_form = request.form['LOGIN']
        mdp = request.form['MDP']

        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT NOM, ROLE, MDP, campagne_id FROM agents WHERE LOGIN = ?", (login_form,))
        row = c.fetchone()

        if row:
            nom_agent, role_agent, hash_en_base, campagne_id = row
            if isinstance(hash_en_base, bytes):
                hash_bytes = hash_en_base
            else:
                hash_bytes = str(hash_en_base).encode('utf-8')

            is_bcrypt = hash_bytes.startswith(b"$2b$") or hash_bytes.startswith(b"$2a$") or hash_bytes.startswith(b"$2y$")
            ok = False
            try:
                if is_bcrypt:
                    ok = bcrypt.checkpw(mdp.encode('utf-8'), hash_bytes)
                else:
                    ok = mdp == (hash_en_base if isinstance(hash_en_base, str) else hash_en_base.decode('utf-8', 'ignore'))
            except Exception:
                ok = False

            if ok:
                session['agent_nom'] = nom_agent
                session['agent_login'] = login_form
                session['agent_role'] = role_agent
                session['campagne_id'] = campagne_id if campagne_id else 1

                flash(f"Bienvenue, {nom_agent} !", "success")
                c.execute(
                    "INSERT INTO journal_connexions (agent_nom, date_connexion, page, type_event) VALUES (?, ?, ?, ?)",
                    (nom_agent, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), '/login', 'connexion')
                )
                conn.commit()
                conn.close()
                return redirect(url_for('index'))
            else:
                conn.close()
                flash("Login ou mot de passe incorrect.", "danger")
        else:
            conn.close()
            flash("Login ou mot de passe incorrect.", "danger")
    return render_template('login.html')

@app.route('/logout')
def logout():
    if 'agent_nom' in session:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute(
            "INSERT INTO journal_connexions (agent_nom, date_connexion, page, type_event) VALUES (?, ?, ?, ?)",
            (session['agent_nom'], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), '/logout', 'deconnexion')
        )
        conn.commit()
        conn.close()
    session.clear()
    flash("Déconnexion réussie.", "info")
    return redirect(url_for('login'))

# ──────────────────────────────────────────────────────────────────────────────
# Formulaires d'ajout
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/', methods=['GET', 'POST'])
def index():
    if 'agent_nom' not in session:
        return redirect(url_for('login'))

    campagne_id = session.get('campagne_id', 1)
    if campagne_id == 2:
        return redirect(url_for('formulaire_valandre'))
    elif campagne_id != 1:
        flash("Accès interdit à ce formulaire.", "danger")
        return redirect(url_for('login'))

    # s'assure que TYPE_OFFRE existe
    conn_check = sqlite3.connect(DB_NAME)
    c_check = conn_check.cursor()
    try:
        c_check.execute("ALTER TABLE clients ADD COLUMN TYPE_OFFRE TEXT DEFAULT ''")
    except sqlite3.OperationalError:
        pass
    conn_check.commit()
    conn_check.close()

    agents = get_agents()
    date_auj = datetime.now().strftime('%Y-%m-%d')

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT nom FROM campagnes WHERE id = ?", (campagne_id,))
    campagne_row = c.fetchone()
    conn.close()
    campagne_nom = campagne_row[0] if campagne_row else "EXOSPHERE_SFR"

    if request.method == 'GET':
        if campagne_nom == "VALANDRE":
            return render_template('formulaire_valandre.html', agents=agents, agent_nom=session['agent_nom'], agent_role=session['agent_role'], date_auj=date_auj)
        else:
            return render_template('formulaire.html', agents=agents, agent_nom=session['agent_nom'], agent_role=session['agent_role'], date_auj=date_auj)

    # POST : enregistrement du client
    if request.method == 'POST':
        campagne_id = session.get('campagne_id', 1)
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("SELECT nom FROM campagnes WHERE id = ?", (campagne_id,))
        campagne_row = c.fetchone()
        conn.close()
        campagne_nom = campagne_row[0] if campagne_row else "EXOSPHERE_SFR"

        if campagne_nom == "VALANDRE":
            # Valandre
            produits = ["STRATO", "LSR", "PRESSE", "ENI", "SERENITY", "PROTEC_ALLIANCE", "WEKIWI"]
            data_dict = {}
            for prod in produits:
                data_dict[f"{prod}_NUM"] = request.form.get(f"{prod}_NUM", "")
                data_dict[f"{prod}_STATUT"] = request.form.get(f"{prod}_STATUT", "")
                data_dict[f"{prod}_REMARQUE"] = request.form.get(f"{prod}_REMARQUE", "")
            data = (
                request.form['DATE_SIGNATURE'],
                request.form['NOM_VENDEUR'],
                request.form['PRENOM_VENDEUR'],
                request.form['TITRE'],
                request.form['NOM_CLIENT'],
                request.form['PRENOM_CLIENT'],
                request.form['TELEPHONE'],
                request.form.get('AGENT', session['agent_nom']),
                request.form.get('EXTRANET', ''),
                *[data_dict[f"{prod}_NUM"] for prod in produits],
                *[data_dict[f"{prod}_STATUT"] for prod in produits],
                *[data_dict[f"{prod}_REMARQUE"] for prod in produits],
                session['agent_nom'],
                session['agent_nom'],
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                campagne_id
            )
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute(f"""
                INSERT INTO clients (
                    DATE_SIGNATURE, NOM_VENDEUR, PRENOM_VENDEUR, TITRE,
                    NOM_CLIENT, PRENOM_CLIENT, TELEPHONE, AGENT, EXTRANET,
                    {', '.join([f"{prod}_NUM" for prod in produits])},
                    {', '.join([f"{prod}_STATUT" for prod in produits])},
                    {', '.join([f"{prod}_REMARQUE" for prod in produits])},
                    CREE_PAR, MODIFIE_PAR, DATE_MODIF, campagne_id
                ) VALUES ({','.join(['?']*(9 + 3*len(produits) + 4))})
            """, data)
            conn.commit()
            conn.close()
            flash("Client VALANDRE ajouté avec succès !", "success")
            return redirect('/dashboard')

        else:
            # SFR
            numero = request.form['TELEPHONE']
            call_id = get_last_aircall_id_by_number(numero)
            deuxieme_adresse = request.form.get('DEUXIEME_ADRESSE', '').strip()
            troisieme_adresse = request.form.get('TROISIEME_ADRESSE', '').strip()
            type_offre = request.form.get('TYPE_OFFRE', '').strip()
            nom_vendeur = request.form.get('NOM_VENDEUR', '').strip()
            prenom_vendeur = request.form.get('PRENOM_VENDEUR', '').strip()
            telephone_vendeur = request.form.get('TELEPHONE_VENDEUR', '').strip()

            if session['agent_role'] == 'admin':
                date_signature = request.form['DATE_SIGNATURE']
                agent_client = request.form['AGENT']
            else:
                date_signature = datetime.now().strftime('%Y-%m-%d')
                agent_client = session['agent_nom']

            date_courante = datetime.strptime(date_signature, "%Y-%m-%d")
            mois_courant = date_courante.month
            annee_courante = date_courante.year

            # Doublons SFR
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("""
                SELECT DATE_SIGNATURE, AGENT, COALESCE(DEUXIEME_ADRESSE,''), COALESCE(TROISIEME_ADRESSE,''), TYPE_OFFRE
                FROM clients
                WHERE TELEPHONE = ? AND STATUT = 'valide'
            """, (numero,))
            doublons = c.fetchall()
            conn.close()

            for date_exist, agent_ancien, adresse2_exist, adresse3_exist, type_offre_exist in doublons:
                try:
                    date_exist_dt = datetime.strptime(date_exist, "%Y-%m-%d")
                    if type_offre.lower() == "mobile":
                        if (type_offre_exist and type_offre_exist.lower() == "mobile" and date_exist_dt.date() == date_courante.date()):
                            flash(f"⚠️ Client Mobile déjà existant saisi par {agent_ancien} pour aujourd'hui.", "danger")
                            return render_template('formulaire.html', agents=agents, agent_nom=session['agent_nom'], agent_role=session['agent_role'], date_auj=date_auj, doublon=True, ancien_agent=agent_ancien, form=request.form)
                        continue
                    if ((not type_offre_exist or type_offre_exist.lower() != "mobile")
                        and date_exist_dt.month == mois_courant
                        and date_exist_dt.year == annee_courante
                        and deuxieme_adresse == "" and troisieme_adresse == ""
                        and adresse2_exist == "" and adresse3_exist == ""):
                        flash(f"⚠️ Client déjà existant saisi par {agent_ancien} pour le même mois.", "danger")
                        return render_template('formulaire.html', agents=agents, agent_nom=session['agent_nom'], agent_role=session['agent_role'], date_auj=date_auj, doublon=True, ancien_agent=agent_ancien, form=request.form)
                except Exception:
                    continue

            # Insert SFR
            champs_insert = [
                'DATE_SIGNATURE','CIVILITE_CLIENT','NOM_CLIENT','PRENOM_CLIENT','TELEPHONE','STATUT','AGENT',
                'DEUXIEME_ADRESSE','TROISIEME_ADRESSE','CREE_PAR','MODIFIE_PAR','DATE_MODIF','campagne_id',
                'NOM_VENDEUR','PRENOM_VENDEUR','TELEPHONE_VENDEUR','TYPE_OFFRE','CALL_ID'
            ]
            data = (
                date_signature, request.form['CIVILITE_CLIENT'], request.form['NOM_CLIENT'], request.form['PRENOM_CLIENT'],
                numero, request.form['STATUT'], agent_client, deuxieme_adresse, troisieme_adresse,
                session['agent_nom'], session['agent_nom'], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), campagne_id,
                nom_vendeur, prenom_vendeur, telephone_vendeur, type_offre, call_id
            )
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            placeholders = ', '.join(['?'] * len(champs_insert))
            sql = f"INSERT INTO clients ({', '.join(champs_insert)}) VALUES ({placeholders})"
            c.execute(sql, data)
            conn.commit()
            conn.close()
            notifier_nouveau_client(request.form['NOM_CLIENT'])
            flash("Client ajouté avec succès !", "success")
            return redirect('/dashboard')

# ──────────────────────────────────────────────────────────────────────────────
# Dashboards
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/dashboard')
def dashboard():
    if 'agent_nom' not in session:
        return redirect(url_for('login'))
    if session.get('agent_role') not in ['admin', 'superviseur'] and session.get('campagne_id') != 1:
        flash("Accès interdit à ce dashboard.", "danger")
        return redirect(url_for('index'))

    recherche = request.args.get('recherche', '').strip()
    statut = request.args.get('statut', '').strip()
    agent = request.args.get('agent', '').strip()
    date_debut = request.args.get('date_debut', '').strip()
    date_fin = request.args.get('date_fin', '').strip()
    type_offre = request.args.get('type_offre', '').strip()

    auj = datetime.now().strftime('%Y-%m-%d')

    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1
    par_page = 20
    offset = (page - 1) * par_page

    if not recherche and not statut and not agent and not date_debut and not date_fin and not type_offre:
        date_debut = date_fin = auj

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("SELECT id FROM campagnes WHERE nom = 'EXOSPHERE_SFR'")
    row = c.fetchone()
    campagne_id = row[0] if row else 1

    # Photo profil
    c.execute("SELECT photo FROM agents WHERE NOM = ?", (session['agent_nom'],))
    photo_row = c.fetchone()
    photo_filename = photo_row[0] if photo_row and photo_row[0] else None
    if photo_filename:
        agent_photo = url_for('static', filename='uploads/' + photo_filename)
    else:
        agent_photo = url_for('static', filename='img/avatar.png')

    c.execute("SELECT NOM FROM agents")
    agents = [row[0] for row in c.fetchall()]

    sql = """SELECT id, DATE_SIGNATURE, CIVILITE_CLIENT, NOM_CLIENT, PRENOM_CLIENT, TELEPHONE, STATUT, AGENT, DEUXIEME_ADRESSE, TROISIEME_ADRESSE, CREE_PAR, MODIFIE_PAR, DATE_MODIF, campagne_id, TITRE, NOM_VENDEUR, PRENOM_VENDEUR, N_CONTRAT, N_REFERENCE, VALIDATION_PRODUIT1, STATUT_PRODUIT1, VALIDATION_PRODUIT2, STATUT_PRODUIT2, VALIDATION_PRODUIT3, STATUT_PRODUIT3, EXTRANET, STRATO_NUM, STRATO_STATUT, STRATO_REMARQUE, LSR_NUM, LSR_STATUT, LSR_REMARQUE, PRESSE_NUM, PRESSE_STATUT, PRESSE_REMARQUE, ENI_NUM, ENI_STATUT, ENI_REMARQUE, SERENITY_NUM, SERENITY_STATUT, SERENITY_REMARQUE, PROTEC_ALLIANCE_NUM, PROTEC_ALLIANCE_STATUT, PROTEC_ALLIANCE_REMARQUE, WEKIWI_NUM, WEKIWI_STATUT, WEKIWI_REMARQUE, TYPE_OFFRE, TELEPHONE_VENDEUR, CALL_ID FROM clients WHERE campagne_id=?"""
    params = [campagne_id]
    if recherche:
        sql += " AND (NOM_CLIENT LIKE ? OR PRENOM_CLIENT LIKE ? OR TELEPHONE LIKE ?)"
        r = f"%{recherche}%"
        params += [r, r, r]
    if statut:
        sql += " AND STATUT=?"
        params.append(statut)
    if agent:
        sql += " AND AGENT=?"
        params.append(agent)
    if type_offre:
        sql += " AND TYPE_OFFRE=?"
        params.append(type_offre)
    if date_debut:
        sql += " AND DATE_SIGNATURE >= ?"
        params.append(date_debut)
    if date_fin:
        sql += " AND DATE_SIGNATURE <= ?"
        params.append(date_fin)

    sql_pagination = sql + " LIMIT ? OFFSET ?"
    params_pagination = params + [par_page, offset]
    c.execute(sql_pagination, params_pagination)
    clients = c.fetchall()

    # Compteurs
    base_sql = "SELECT COUNT(*) FROM clients WHERE campagne_id = ?"
    params_count = [campagne_id]
    if recherche:
        base_sql += " AND (NOM_CLIENT LIKE ? OR PRENOM_CLIENT LIKE ? OR TELEPHONE LIKE ?)"
        r = f"%{recherche}%"
        params_count += [r, r, r]
    if statut:
        base_sql += " AND STATUT=?"
        params_count.append(statut)
    if agent:
        base_sql += " AND AGENT=?"
        params_count.append(agent)
    if type_offre:
        base_sql += " AND TYPE_OFFRE=?"
        params_count.append(type_offre)
    if date_debut:
        base_sql += " AND DATE_SIGNATURE >= ?"
        params_count.append(date_debut)
    if date_fin:
        base_sql += " AND DATE_SIGNATURE <= ?"
        params_count.append(date_fin)

    def get_count_for_statut(statut_value):
        sql_count = base_sql + " AND STATUT=?"
        c.execute(sql_count, params_count + [statut_value])
        result = c.fetchone()
        return result[0] if result and result[0] is not None else 0

    count_valide = get_count_for_statut('valide')
    count_non_valide = get_count_for_statut('non valide')

    # Total pour pagination
    count_sql = "SELECT COUNT(*) FROM clients WHERE campagne_id = ?"
    count_params = [campagne_id]
    if recherche:
        count_sql += " AND (NOM_CLIENT LIKE ? OR PRENOM_CLIENT LIKE ? OR TELEPHONE LIKE ?)"
        r = f"%{recherche}%"
        count_params += [r, r, r]
    if statut:
        count_sql += " AND STATUT=?"
        count_params.append(statut)
    if agent:
        count_sql += " AND AGENT=?"
        count_params.append(agent)
    if type_offre:
        count_sql += " AND TYPE_OFFRE=?"
        count_params.append(type_offre)
    if date_debut:
        count_sql += " AND DATE_SIGNATURE >= ?"
        count_params.append(date_debut)
    if date_fin:
        count_sql += " AND DATE_SIGNATURE <= ?"
        count_params.append(date_fin)

    c.execute(count_sql, count_params)
    total_clients = c.fetchone()[0]
    total_pages = (total_clients + par_page - 1) // par_page

    conn.close()

    class Pagination:
        def __init__(self, page, total_pages):
            self.page = page
            self.total_pages = total_pages
        @property
        def has_prev(self):
            return self.page > 1
        @property
        def has_next(self):
            return self.page < self.total_pages
        @property
        def prev_num(self):
            return self.page - 1
        @property
        def next_num(self):
            return self.page + 1
        def iter_pages(self):
            left = max(1, self.page - 2)
            right = min(self.total_pages, self.page + 2)
            return range(left, right + 1)

    pagination = Pagination(page, total_pages)

    return render_template(
        'dashboard.html',
        clients=clients,
        agents=agents,
        count_valide=count_valide,
        count_non_valide=count_non_valide,
        auj=auj,
        agent_photo=agent_photo,
        pagination=pagination,
        type_offre=type_offre
    )

@app.route('/dashboard_valandre')
def dashboard_valandre():
    if 'agent_nom' not in session:
        return redirect(url_for('login'))
    if session.get('agent_role') not in ['admin', 'superviseur'] and session.get('campagne_id') != 2:
        flash("Accès interdit à ce dashboard.", "danger")
        return redirect(url_for('index'))

    try:
        page = int(request.args.get('page', 1))
        if page < 1:
            page = 1
    except ValueError:
        page = 1
    par_page = 20
    offset = (page - 1) * par_page

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("SELECT photo FROM agents WHERE NOM = ?", (session['agent_nom'],))
    photo_row = c.fetchone()
    photo_filename = photo_row[0] if photo_row and photo_row[0] else None
    agent_photo = url_for('static', filename='uploads/' + photo_filename) if photo_filename else url_for('static', filename='img/avatar.png')

    c.execute("SELECT id FROM campagnes WHERE nom = 'VALANDRE'")
    campagne_row = c.fetchone()
    campagne_id = campagne_row[0] if campagne_row else 2

    agents = get_agents()
    auj = datetime.now().strftime('%Y-%m-%d')

    produits = ["STRATO", "LSR", "PRESSE", "ENI", "SERENITY", "PROTEC_ALLIANCE", "WEKIWI"]
    champs = ['id','DATE_SIGNATURE','NOM_VENDEUR','PRENOM_VENDEUR','TITRE','NOM_CLIENT','PRENOM_CLIENT','TELEPHONE']
    for prod in produits:
        champs += [f'{prod}_NUM', f'{prod}_STATUT', f'{prod}_REMARQUE']
    champs += ['AGENT','EXTRANET']

    filters = []
    params = [campagne_id]
    telephone = request.args.get('telephone', '').strip()
    date_debut = request.args.get('date_debut', '').strip()
    date_fin = request.args.get('date_fin', '').strip()

    if telephone:
        filters.append("TELEPHONE LIKE ?")
        params.append(f"%{telephone}%")

    if not date_debut and not date_fin and not telephone:
        date_debut = date_fin = auj

    if date_debut and date_fin:
        filters.append("DATE_SIGNATURE BETWEEN ? AND ?")
        params.extend([date_debut, date_fin])
    elif date_debut:
        filters.append("DATE_SIGNATURE >= ?")
        params.append(date_debut)
    elif date_fin:
        filters.append("DATE_SIGNATURE <= ?")
        params.append(date_fin)

    sql = f"SELECT {', '.join(champs)} FROM clients WHERE campagne_id=?"
    if filters:
        sql += " AND " + " AND ".join(filters)
    sql_pagination = sql + " ORDER BY DATE_SIGNATURE DESC LIMIT ? OFFSET ?"
    params_pagination = params + [par_page, offset]
    c.execute(sql_pagination, tuple(params_pagination))
    rows = c.fetchall()
    clients = [dict(zip(champs, row)) for row in rows]

    count_sql = "SELECT COUNT(*) FROM clients WHERE campagne_id=?"
    count_params = [campagne_id]
    if filters:
        count_sql += " AND " + " AND ".join(filters)
        count_params += params[1:]
    c.execute(count_sql, tuple(count_params))
    total_clients = c.fetchone()[0]
    total_pages = (total_clients + par_page - 1) // par_page

    c.execute("""
        SELECT COUNT(*) FROM clients
        WHERE campagne_id = ? AND DATE_SIGNATURE = ?
        AND (STRATO_STATUT='VALIDÉ' OR LSR_STATUT='VALIDÉ' OR PRESSE_STATUT='VALIDÉ' OR ENI_STATUT='VALIDÉ' OR SERENITY_STATUT='VALIDÉ' OR PROTEC_ALLIANCE_STATUT='VALIDÉ' OR WEKIWI_STATUT='VALIDÉ')
    """, (campagne_id, auj))
    count_valide = c.fetchone()[0]

    c.execute("""
        SELECT COUNT(*) FROM clients
        WHERE campagne_id = ? AND DATE_SIGNATURE = ?
        AND (STRATO_STATUT='REFUSÉ' OR LSR_STATUT='REFUSÉ' OR PRESSE_STATUT='REFUSÉ' OR ENI_STATUT='REFUSÉ' OR SERENITY_STATUT='REFUSÉ' OR PROTEC_ALLIANCE_STATUT='REFUSÉ' OR WEKIWI_STATUT='REFUSÉ')
        AND NOT (STRATO_STATUT='VALIDÉ' OR LSR_STATUT='VALIDÉ' OR PRESSE_STATUT='VALIDÉ' OR ENI_STATUT='VALIDÉ' OR SERENITY_STATUT='VALIDÉ' OR PROTEC_ALLIANCE_STATUT='VALIDÉ' OR WEKIWI_STATUT='VALIDÉ')
    """, (campagne_id, auj))
    count_refuse = c.fetchone()[0]

    conn.close()

    class Pagination:
        def __init__(self, page, total_pages):
            self.page = page
            self.total_pages = total_pages
        @property
        def has_prev(self): return self.page > 1
        @property
        def has_next(self): return self.page < self.total_pages
        @property
        def prev_num(self): return self.page - 1
        @property
        def next_num(self): return self.page + 1
        def iter_pages(self):
            left = max(1, self.page - 2)
            right = min(self.total_pages, self.page + 2)
            return range(left, right + 1)

    pagination = Pagination(page, total_pages)

    return render_template('dashboard_valandre.html',
                           clients=clients, agents=agents, auj=auj, agent_photo=agent_photo,
                           date_debut=date_debut, date_fin=date_fin,
                           count_valide=count_valide, count_refuse=count_refuse,
                           pagination=pagination)

# ──────────────────────────────────────────────────────────────────────────────
# Formulaire Valandre
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/formulaire_valandre', methods=['GET', 'POST'])
def formulaire_valandre():
    if 'agent_nom' not in session:
        return redirect(url_for('login'))
    if session.get('agent_role') not in ['admin', 'superviseur'] and session.get('campagne_id') != 2:
        flash("Accès interdit à ce formulaire.", "danger")
        return redirect(url_for('index'))

    if request.method == 'POST':
        nom_client = request.form.get("NOM_CLIENT")
        prenom_client = request.form.get("PRENOM_CLIENT")
        telephone = request.form.get("TELEPHONE")
        nom_vendeur = request.form.get("NOM_VENDEUR")
        prenom_vendeur = request.form.get("PRENOM_VENDEUR")
        titre = request.form.get("TITRE")
        produits_sel = request.form.getlist("produits[]")

        data_produits = {}
        for produit in produits_sel:
            num = request.form.get(f"{produit}_NUM")
            statut = request.form.get(f"{produit}_STATUT")
            remarque = request.form.get(f"{produit}_REMARQUE")
            if not num or not statut:
                flash(f"Veuillez remplir le numéro et le statut pour le produit {produit.replace('_',' ')}.", "danger")
                return redirect(url_for("formulaire_valandre"))
            data_produits[produit] = {"num": num, "statut": statut, "remarque": remarque or ""}

        colonnes_fixes = ["DATE_SIGNATURE", "NOM_VENDEUR", "PRENOM_VENDEUR", "TITRE", "NOM_CLIENT", "PRENOM_CLIENT", "TELEPHONE", "AGENT", "EXTRANET"]
        produits = ["STRATO","LSR","PRESSE","ENI","SERENITY","PROTEC_ALLIANCE","WEKIWI"]
        colonnes_produits = []
        for prod in produits:
            colonnes_produits += [f"{prod}_NUM", f"{prod}_STATUT", f"{prod}_REMARQUE"]
        colonnes = colonnes_fixes + colonnes_produits

        # (Bloc export rapide s'il est déclenché via ce POST)
        conn = sqlite3.connect(DB_NAME)
        df = pd.read_sql_query("SELECT * FROM clients", conn)
        conn.close()
        df = df.reindex(columns=colonnes, fill_value="")
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Clients")
        output.seek(0)
        return send_file(output, download_name="export_clients.xlsx", as_attachment=True)

    agents = get_agents()
    date_auj = datetime.now().strftime('%Y-%m-%d')
    return render_template('formulaire_valandre.html',
                           agents=agents, agent_nom=session['agent_nom'],
                           agent_role=session['agent_role'], date_auj=date_auj)

# ──────────────────────────────────────────────────────────────────────────────
# Paramètres / Agents
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/parametres', methods=['GET', 'POST'])
def parametres():
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        flash("Accès réservé à l'administration/supervision.", "danger")
        return redirect(url_for('index'))
    roles = ["agent", "admin", "superviseur"]
    if request.method == 'POST':
        nom = request.form['NOM']
        login = request.form['LOGIN']
        mdp = request.form['MDP']
        role = request.form['ROLE']

        photo_filename = None
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename != '':
                filename = secure_filename(file.filename)
                if not allowed_file(filename):
                    flash("Format de fichier non autorisé. Utilise JPG, JPEG ou PNG.", "danger")
                    return redirect(request.url)
                if not file_size_okay(file):
                    flash("Le fichier est trop lourd. Taille max : 5 Mo.", "danger")
                    return redirect(request.url)
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                photo_filename = filename

        hashed_mdp = bcrypt.hashpw(mdp.encode('utf-8'), bcrypt.gensalt())
        try:
            conn = sqlite3.connect(DB_NAME)
            c = conn.cursor()
            c.execute("INSERT INTO agents (NOM, LOGIN, MDP, ROLE, photo) VALUES (?, ?, ?, ?, ?)",
                      (nom, login, hashed_mdp, role, photo_filename))
            conn.commit()
            conn.close()
            flash("Nouvel agent créé avec succès !", "success")
        except sqlite3.IntegrityError:
            flash("Nom ou login déjà utilisé !", "danger")
        return redirect('/parametres')

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        SELECT a.id, a.NOM, a.LOGIN, a.ROLE, c.nom
        FROM agents a
        LEFT JOIN campagnes c ON a.campagne_id = c.id
    """)
    agents = c.fetchall()
    conn.close()
    campagnes = get_campagnes()
    return render_template('parametres.html', agents=agents, roles=roles, campagnes=campagnes)

@app.route('/profil', methods=['GET', 'POST'])
def profil_agent():
    if 'agent_nom' not in session:
        return redirect(url_for('login'))

    agent_nom = session['agent_nom']
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id, NOM, photo FROM agents WHERE NOM = ?", (agent_nom,))
    agent = c.fetchone()
    if not agent:
        conn.close()
        flash("Agent introuvable.", "danger")
        return redirect(url_for('dashboard'))

    agent_id, agent_nom_db, agent_photo = agent
    photo_url = url_for('static', filename='uploads/' + agent_photo) if agent_photo else url_for('static', filename='img/avatar.png')

    c.execute("SELECT COUNT(*) FROM clients WHERE AGENT = ? AND STATUT = 'valide'", (agent_nom,))
    clients_valides = c.fetchone()[0] or 0
    c.execute("SELECT COUNT(*) FROM clients WHERE AGENT = ?", (agent_nom,))
    total_clients = c.fetchone()[0] or 0

    c.execute("""
        SELECT AGENT, COUNT(*) as nb_valides
        FROM clients
        WHERE STATUT = 'valide'
        GROUP BY AGENT
        ORDER BY nb_valides DESC
    """)
    classement = c.fetchall()
    position = "-"
    for idx, (nom, nb) in enumerate(classement, 1):
        if nom == agent_nom:
            position = idx
            break

    if request.method == 'POST':
        if 'photo' in request.files:
            file = request.files['photo']
            if file and file.filename != '':
                filename = secure_filename(f"{agent_id}_{file.filename}")
                file.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
                c.execute("UPDATE agents SET photo=? WHERE id=?", (filename, agent_id))
                conn.commit()
                flash("Photo de profil mise à jour avec succès !", "success")
                conn.close()
                return redirect(url_for('profil_agent'))
        flash("Aucune photo sélectionnée.", "warning")
    conn.close()
    return render_template('profil_agent.html', agent_nom=agent_nom, photo_url=photo_url,
                           clients_valides=clients_valides, total_clients=total_clients, position=position)

@app.route('/modifier_agent/<int:agent_id>', methods=['GET', 'POST'])
def modifier_agent(agent_id):
    if 'agent_nom' not in session:
        return redirect(url_for('login'))
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        flash("Accès réservé à l'administration/supervision.", "danger")
        return redirect(url_for('index'))

    roles = ["agent", "admin", "superviseur"]
    campagnes = get_campagnes()

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id, NOM, LOGIN, ROLE, campagne_id FROM agents WHERE id = ?", (agent_id,))
    agent = c.fetchone()

    if request.method == 'POST':
        # Récupération des champs du formulaire (avec valeurs par défaut sûres)
        nom = request.form.get('NOM', '').strip()
        login = request.form.get('LOGIN', '').strip()
        mdp = request.form.get('MDP', '')
        role = request.form.get('ROLE', 'agent').strip()

        # ⚠️ CAMPAGNE_ID peut ne pas être envoyé par le formulaire de modif.
        # Si présent, on le prend ; sinon, on conserve la valeur actuelle en base.
        campagne_id_str = request.form.get('CAMPAGNE_ID', '').strip()
        if campagne_id_str == '':
            # Garder la valeur existante
            c.execute("SELECT campagne_id FROM agents WHERE id=?", (agent_id,))
            row = c.fetchone()
            campagne_id = row[0] if row else None
        else:
            # Convertir proprement en int, sinon mettre à None
            try:
                campagne_id = int(campagne_id_str)
            except ValueError:
                campagne_id = None

        # ✅ Mise à jour : si mdp non vide -> on hash, sinon on ne touche pas à MDP
        if mdp.strip():
            hashed_mdp = bcrypt.hashpw(mdp.encode('utf-8'), bcrypt.gensalt())
            c.execute(
                "UPDATE agents SET NOM=?, LOGIN=?, MDP=?, ROLE=?, campagne_id=? WHERE id=?",
                (nom, login, hashed_mdp, role, campagne_id, agent_id)
            )
        else:
            c.execute(
                "UPDATE agents SET NOM=?, LOGIN=?, ROLE=?, campagne_id=? WHERE id=?",
                (nom, login, role, campagne_id, agent_id)
            )

        conn.commit()
        conn.close()
        flash("Agent modifié avec succès.", "success")
        return redirect(url_for('parametres'))

    conn.close()
    return render_template('modifier_agent.html', agent=agent, roles=roles, campagnes=campagnes)

@app.route('/supprimer_agent/<int:agent_id>', methods=['POST'])
def supprimer_agent(agent_id):
    if 'agent_nom' not in session:
        return redirect(url_for('login'))
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        flash("Accès réservé à l'administration/supervision.", "danger")
        return redirect(url_for('index'))

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT COUNT(*) FROM agents")
    count = c.fetchone()[0]
    if count <= 1:
        flash("Impossible de supprimer le dernier agent.", "danger")
    else:
        c.execute("DELETE FROM agents WHERE id=?", (agent_id,))
        conn.commit()
        flash("Agent supprimé avec succès.", "success")
    conn.close()
    return redirect(url_for('parametres'))

@app.route('/supprimer_client/<int:client_id>', methods=['POST'])
def supprimer_client(client_id):
    if 'agent_nom' not in session:
        return redirect(url_for('login'))
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # --- CONTRÔLE D'AUTORISATION ---
    role = session.get('agent_role')
    agent_session = session.get('agent_nom')

    c.execute("SELECT AGENT FROM clients WHERE id=?", (client_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        flash("Client introuvable.", "danger")
        return redirect(url_for('dashboard'))

    owner = row[0]

    if role not in ("admin", "superviseur") and owner != agent_session:
        conn.close()
        flash("Accès refusé : vous ne pouvez supprimer que vos propres fiches.", "danger")
        return redirect(url_for('dashboard'))

    c.execute("DELETE FROM clients WHERE id=?", (client_id,))
    conn.commit()
    conn.close()
    flash("Client supprimé.", "info")
    return redirect(url_for('dashboard'))

@app.route('/modifier_client/<int:client_id>', methods=['GET', 'POST'])
def modifier_client(client_id):
    if 'agent_nom' not in session:
        return redirect(url_for('login'))
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    # --- CONTRÔLE D'AUTORISATION ---
    role = session.get('agent_role')
    agent_session = session.get('agent_nom')

    # On récupère le propriétaire de la fiche (colonne AGENT)
    c.execute("SELECT AGENT FROM clients WHERE id= ?", (client_id,))
    row = c.fetchone()
    if not row:
        conn.close()
        flash("Client introuvable.", "danger")
        return redirect(url_for('dashboard'))

    owner = row[0]

    # Seuls admin/superviseur OU le propriétaire peuvent modifier
    if role not in ("admin", "superviseur") and owner != agent_session:
        conn.close()
        flash("Accès refusé : vous ne pouvez modifier que vos propres fiches.", "danger")
        return redirect(url_for('dashboard'))

    c.execute("SELECT campagne_id FROM clients WHERE id= ?", (client_id,))
    campagne_id_row = c.fetchone()
    campagne_id = campagne_id_row[0] if campagne_id_row else 1
    c.execute("SELECT nom FROM campagnes WHERE id= ?", (campagne_id,))
    campagne_nom_row = c.fetchone()
    campagne_nom = campagne_nom_row[0] if campagne_nom_row else "EXOSPHERE_SFR"

    if request.method == 'POST':
        if campagne_nom == "VALANDRE":
            produits = ["STRATO", "LSR", "PRESSE", "ENI", "SERENITY", "PROTEC_ALLIANCE", "WEKIWI"]
            champs = ['DATE_SIGNATURE','NOM_VENDEUR','PRENOM_VENDEUR','TITRE','NOM_CLIENT','PRENOM_CLIENT','TELEPHONE']
            for prod in produits:
                champs += [f"{prod}_NUM", f"{prod}_STATUT", f"{prod}_REMARQUE"]
            champs += ['EXTRANET','AGENT']

            c.execute(f"SELECT {', '.join(champs)} FROM clients WHERE id= ?", (client_id,))
            ancien = c.fetchone()

            nouveaux = (
                request.form['DATE_SIGNATURE'], request.form['NOM_VENDEUR'], request.form['PRENOM_VENDEUR'], request.form['TITRE'],
                request.form['NOM_CLIENT'], request.form['PRENOM_CLIENT'], request.form['TELEPHONE'],
            )
            for prod in produits:
                nouveaux += (request.form.get(f"{prod}_NUM",""), request.form.get(f"{prod}_STATUT",""), request.form.get(f"{prod}_REMARQUE",""))
            nouveaux += (request.form.get('EXTRANET',''), request.form['AGENT'])

            for i in range(len(champs)):
                if str(ancien[i]) != str(nouveaux[i]):
                    c.execute("""
                        INSERT INTO historique_clients (client_id, date_modif, agent, champ_modifie, ancienne_valeur, nouvelle_valeur)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (client_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), session['agent_nom'], champs[i], ancien[i], nouveaux[i]))

            update_fields = ", ".join([f"{champs[i]}=?" for i in range(len(champs))])
            data = nouveaux + (session['agent_nom'], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), client_id)
            c.execute(f"UPDATE clients SET {update_fields}, MODIFIE_PAR=?, DATE_MODIF=? WHERE id= ?", data)
            conn.commit()
            conn.close()
            notifier_nouveau_client(request.form['NOM_CLIENT'])
            flash("Client VALANDRE modifié avec succès.", "success")
            return redirect(url_for('dashboard_valandre'))

        else:
            c.execute("SELECT DATE_SIGNATURE, CIVILITE_CLIENT, NOM_CLIENT, PRENOM_CLIENT, TELEPHONE, STATUT, AGENT, DEUXIEME_ADRESSE, TROISIEME_ADRESSE FROM clients WHERE id= ?", (client_id,))
            ancien = c.fetchone()
            nouveaux = (
                request.form['DATE_SIGNATURE'], request.form['CIVILITE_CLIENT'], request.form['NOM_CLIENT'], request.form['PRENOM_CLIENT'],
                request.form['TELEPHONE'], request.form['STATUT'], request.form['AGENT'],
                request.form.get('DEUXIEME_ADRESSE',''), request.form.get('TROISIEME_ADRESSE','')
            )
            champs = ['DATE_SIGNATURE','CIVILITE_CLIENT','NOM_CLIENT','PRENOM_CLIENT','TELEPHONE','STATUT','AGENT','DEUXIEME_ADRESSE','TROISIEME_ADRESSE']

            for i in range(len(champs)):
                if str(ancien[i]) != str(nouveaux[i]):
                    c.execute("""
                        INSERT INTO historique_clients (client_id, date_modif, agent, champ_modifie, ancienne_valeur, nouvelle_valeur)
                        VALUES (?, ?, ?, ?, ?, ?)
                    """, (client_id, datetime.now().strftime("%Y-%m-%d %H:%M:%S"), session['agent_nom'], champs[i], ancien[i], nouveaux[i]))

            data = nouveaux + (session['agent_nom'], datetime.now().strftime("%Y-%m-%d %H:%M:%S"), client_id)
            c.execute("""
                UPDATE clients
                SET DATE_SIGNATURE=?, CIVILITE_CLIENT=?, NOM_CLIENT=?, PRENOM_CLIENT=?, TELEPHONE=?, STATUT=?, AGENT=?, DEUXIEME_ADRESSE=?, TROISIEME_ADRESSE=?,
                    MODIFIE_PAR=?, DATE_MODIF=?
                WHERE id= ?
            """, data)
            conn.commit()
            conn.close()
            notifier_nouveau_client(request.form['NOM_CLIENT'])
            flash("Client modifié avec succès.", "success")
            return redirect(url_for('dashboard_valandre'))

    if campagne_nom == "VALANDRE":
        produits = ["STRATO", "LSR", "PRESSE", "ENI", "SERENITY", "PROTEC_ALLIANCE", "WEKIWI"]
        champs = ['id','DATE_SIGNATURE','NOM_VENDEUR','PRENOM_VENDEUR','TITRE','NOM_CLIENT','PRENOM_CLIENT','TELEPHONE']
        for prod in produits:
            champs += [f'{prod}_NUM', f'{prod}_STATUT', f'{prod}_REMARQUE']
        champs += ['EXTRANET','AGENT']
        c.execute(f"SELECT {', '.join(champs)} FROM clients WHERE id= ?", (client_id,))
        client_row = c.fetchone()
        client = dict(zip(champs, client_row))
        agents = get_agents()
        conn.close()
        return render_template('modifier_client_valandre.html', client=client, agents=agents)
    else:
        c.execute("SELECT id, DATE_SIGNATURE, CIVILITE_CLIENT, NOM_CLIENT, PRENOM_CLIENT, TELEPHONE, STATUT, AGENT, DEUXIEME_ADRESSE, TROISIEME_ADRESSE FROM clients WHERE id= ?", (client_id,))
        client = c.fetchone()
        c.execute("SELECT NOM FROM agents")
        agents = [row[0] for row in c.fetchall()]
        conn.close()
        return render_template('modifier_client.html', client=client, agents=agents)

# ──────────────────────────────────────────────────────────────────────────────
# Historique client
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/historique_client/<int:client_id>')
def historique_client(client_id):
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        SELECT date_modif, agent, champ_modifie, ancienne_valeur, nouvelle_valeur
        FROM historique_clients
        WHERE client_id=?
        ORDER BY date_modif DESC
    """, (client_id,))
    historique = c.fetchall()
    conn.close()
    return render_template('historique_client.html', historique=historique, client_id=client_id)

# ──────────────────────────────────────────────────────────────────────────────
# Exports
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/export_excel_valandre')
def export_excel_valandre():
    if 'agent_nom' not in session:
        return redirect(url_for('login'))

    def clean_date(dt):
        if not dt:
            return ""
        dt = dt.strip()
        if "-" in dt:
            return dt
        elif "/" in dt:
            try:
                return datetime.strptime(dt, "%d/%m/%Y").strftime("%Y-%m-%d")
            except:
                return dt
        return dt

    date_debut = clean_date(request.args.get('date_debut', '').strip())
    date_fin = clean_date(request.args.get('date_fin', '').strip())
    telephone = request.args.get('telephone', '').strip()
    agent = request.args.get('agent', '').strip()
    statut = request.args.get('statut', '').strip()

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id FROM campagnes WHERE nom = 'VALANDRE'")
    campagne_row = c.fetchone()
    campagne_id = campagne_row[0] if campagne_row else 2

    produits = ["STRATO", "LSR", "PRESSE", "ENI", "SERENITY", "PROTEC_ALLIANCE", "WEKIWI"]
    select_cols = ["DATE_SIGNATURE", "NOM_VENDEUR", "PRENOM_VENDEUR", "TITRE", "NOM_CLIENT", "PRENOM_CLIENT", "TELEPHONE"]
    for prod in produits:
        select_cols += [f"{prod}_NUM", f"{prod}_STATUT", f"{prod}_REMARQUE"]
    select_cols += ["EXTRANET", "AGENT"]

    sql = f"SELECT {','.join(select_cols)} FROM clients WHERE campagne_id=?"
    params = [campagne_id]
    if date_debut and date_fin:
        sql += " AND DATE_SIGNATURE BETWEEN ? AND ?"
        params += [date_debut, date_fin]
    elif date_debut:
        sql += " AND DATE_SIGNATURE >= ?"
        params.append(date_debut)
    elif date_fin:
        sql += " AND DATE_SIGNATURE <= ?"
        params.append(date_fin)
    if telephone:
        sql += " AND TELEPHONE LIKE ?"
        params.append(f"%{telephone}%")
    if agent:
        sql += " AND AGENT=?"
        params.append(agent)
    if statut:
        sql += " AND STATUT=?"
        params.append(statut)
    sql += " ORDER BY DATE_SIGNATURE DESC"

    c.execute(sql, params)
    rows = c.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active

    header1 = ["DATE DE SIGNATURE", "NOM VENDEUR", "PRENOM VENDEUR", "TITRE", "NOM CLIENT", "PRENOM CLIENT", "TÉLÉPHONE"]
    for prod in produits:
        header1.extend([f"VALIDATION {prod}"] * 3)
    header1.extend(["EXTRANET", "AGENT"])
    ws.append(header1)

    header2 = ["", "", "", "", "", "", ""]
    for _ in produits:
        header2.extend(["N° CONTRAT/RÉF", "STATUT", "REMARQUE"])
    header2.extend(["", ""])
    ws.append(header2)

    col = 1
    for _ in range(7):
        ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)
        col += 1
    for _ in produits:
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+2)
        col += 3
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col); col += 1
    ws.merge_cells(start_row=1, start_column=col, end_row=2, end_column=col)

    for row in rows:
        ws.append(list(row))

    for cell in ws["1:1"]:
        cell.alignment = Alignment(horizontal='center', vertical='center')
    for cell in ws["2:2"]:
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for idx, column_cells in enumerate(ws.columns, 1):
        max_length = 0
        for cell in column_cells:
            if cell.value:
                try:
                    max_length = max(max_length, len(str(cell.value)))
                except:
                    continue
        col_letter = get_column_letter(idx)
        ws.column_dimensions[col_letter].width = max_length + 2

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        wb.save(tmp.name)
        tmp_path = tmp.name

    response = send_file(tmp_path, as_attachment=True)
    try:
        os.remove(tmp_path)
    except Exception:
        pass

    return response

@app.route('/export_excel_sfr')
def export_excel_sfr():
    if 'agent_nom' not in session:
        return redirect(url_for('login'))

    date_debut = request.args.get('date_debut', '').strip()
    date_fin = request.args.get('date_fin', '').strip()
    auj = datetime.now().strftime('%Y-%m-%d')
    if not date_debut and not date_fin:
        date_debut = date_fin = auj

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT id FROM campagnes WHERE nom = 'EXOSPHERE_SFR'")
    campagne_row = c.fetchone()
    campagne_id = campagne_row[0] if campagne_row else 1

    sql = """
        SELECT DATE_SIGNATURE, CIVILITE_CLIENT, NOM_CLIENT, PRENOM_CLIENT, TELEPHONE, STATUT, AGENT, DEUXIEME_ADRESSE
        FROM clients
        WHERE campagne_id=?
    """
    params = [campagne_id]
    if date_debut and date_fin:
        sql += " AND DATE_SIGNATURE BETWEEN ? AND ?"
        params += [date_debut, date_fin]
    elif date_debut:
        sql += " AND DATE_SIGNATURE >= ?"
        params.append(date_debut)
    elif date_fin:
        sql += " AND DATE_SIGNATURE <= ?"
        params.append(date_fin)
    sql += " ORDER BY DATE_SIGNATURE DESC"

    c.execute(sql, params)
    rows = c.fetchall()
    conn.close()

    columns = ["DATE_SIGNATURE","CIVILITE_CLIENT","NOM_CLIENT","PRENOM_CLIENT","TELEPHONE","STATUT","AGENT","DEUXIEME_ADRESSE"]
    df = pd.DataFrame(rows, columns=columns)

    with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
        df.to_excel(tmp.name, index=False)
        tmp_path = tmp.name

    response = send_file(tmp_path, as_attachment=True)
    try:
        os.remove(tmp_path)
    except Exception:
        pass

    return response

# ──────────────────────────────────────────────────────────────────────────────
# Journal / Présence / Live
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/journal')
def journal():
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        flash("Accès réservé à l'administration/supervision.", "danger")
        return redirect(url_for('dashboard'))

    agent = request.args.get('agent', '').strip()
    date_debut = request.args.get('date_debut', '').strip()
    date_fin = request.args.get('date_fin', '').strip()

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT DISTINCT agent_nom FROM journal_connexions")
    agents = [row[0] for row in c.fetchall()]

    auj = datetime.now().strftime('%Y-%m-%d')
    if not date_debut and not date_fin:
        date_debut = date_fin = auj

    sql = "SELECT agent_nom, date_connexion, page, type_event FROM journal_connexions WHERE 1=1"
    params = []
    if agent:
        sql += " AND agent_nom = ?"; params.append(agent)
    if date_debut:
        sql += " AND date_connexion >= ?"; params.append(date_debut + " 00:00:00")
    if date_fin:
        sql += " AND date_connexion <= ?"; params.append(date_fin + " 23:59:59")
    sql += " ORDER BY agent_nom, date_connexion ASC"
    c.execute(sql, params)
    all_logs = c.fetchall()
    conn.close()

    from collections import defaultdict
    logs_by_agent_by_day = defaultdict(lambda: defaultdict(list))
    for log in all_logs:
        agent_name = log[0]
        day_str = log[1][:10]
        logs_by_agent_by_day[agent_name][day_str].append(log)

    filtered_logs = []
    for agent_name, days in logs_by_agent_by_day.items():
        for day, logs in days.items():
            first_conn = None
            last_deconn = None
            for l in logs:
                if l[3] == 'connexion' and not first_conn:
                    first_conn = l
                if l[3] in ['connexion', 'deconnexion']:
                    last_deconn = l
            if first_conn:
                filtered_logs.append(first_conn)
            if last_deconn and last_deconn != first_conn:
                filtered_logs.append(last_deconn)

    filtered_logs.sort(key=lambda x: (x[0], x[1]))

    return render_template('journal.html', logs=filtered_logs, agents=agents)

@app.route('/export_journal')
def export_journal():
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        flash("Accès réservé à l'administration/supervision.", "danger")
        return redirect(url_for('dashboard'))

    agent = request.args.get('agent', '').strip()
    date_debut = request.args.get('date_debut', '').strip()
    date_fin = request.args.get('date_fin', '').strip()

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    sql = "SELECT agent_nom, date_connexion, page, type_event FROM journal_connexions WHERE 1=1"
    params = []
    if agent:
        sql += " AND agent_nom = ?"; params.append(agent)
    if date_debut:
        sql += " AND date_connexion >= ?"; params.append(date_debut + " 00:00:00")
    if date_fin:
        sql += " AND date_connexion <= ?"; params.append(date_fin + " 23:59:59")
    sql += " ORDER BY date_connexion DESC"
    c.execute(sql, params)
    rows = c.fetchall()
    conn.close()

    columns = ["Agent","Date/Heure","Page","Événement"]
    df = pd.DataFrame(rows, columns=columns)
    file_path = "export_journal.xlsx"
    df.to_excel(file_path, index=False)
    return send_file(file_path, as_attachment=True)

@app.route('/journal_presence')
def journal_presence():
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        flash("Accès réservé à l'administration/supervision.", "danger")
        return redirect(url_for('dashboard'))

    agent = request.args.get('agent', '').strip()
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')

    from datetime import datetime as dt, timedelta
    auj_str = dt.now().strftime('%Y-%m-%d')

    if not date_debut and not date_fin:
        date_debut = date_fin = auj_str
    elif not date_debut:
        date_debut = date_fin
    elif not date_fin:
        date_fin = date_debut

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT DISTINCT agent_nom FROM journal_connexions")
    agents = [row[0] for row in c.fetchall()]

    d1 = dt.strptime(date_debut, '%Y-%m-%d')
    d2 = dt.strptime(date_fin, '%Y-%m-%d')
    jours = []
    d = d1
    while d <= d2:
        jours.append(d.strftime('%Y-%m-%d'))
        d += timedelta(days=1)

    agents_a_afficher = [agent] if agent else agents

    tableau = []
    for ag in agents_a_afficher:
        for jour in jours:
            c.execute("""
                SELECT MIN(date_connexion), MAX(date_connexion)
                FROM journal_connexions
                WHERE agent_nom=? AND date_connexion>=? AND date_connexion<=? AND (type_event='connexion' OR type_event='deconnexion')
            """, (ag, jour+" 00:00:00", jour+" 23:59:59"))
            entree, sortie = c.fetchone()
            heure_entree = entree[11:19] if entree else ''
            heure_sortie = sortie[11:19] if sortie else ''
            if entree and sortie:
                dt1 = dt.strptime(entree, "%Y-%m-%d %H:%M:%S")
                dt2 = dt.strptime(sortie, "%Y-%m-%d %H:%M:%S")
                duree = dt2 - dt1 if dt2 > dt1 else timedelta()
                h = int(duree.total_seconds() // 3600)
                m = int((duree.total_seconds() % 3600) // 60)
                duree_txt = f"{h:02d}:{m:02d}"
            else:
                duree_txt = ''
            tableau.append([ag, jour, heure_entree, heure_sortie, duree_txt])

    conn.close()
    return render_template('journal_presence.html', tableau=tableau, jours=jours, agents=agents,
                           date_debut=date_debut, date_fin=date_fin, agent_selected=agent)

@app.route('/export_presence')
def export_presence():
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        flash("Accès réservé à l'administration/supervision.", "danger")
        return redirect(url_for('dashboard'))

    from datetime import datetime as dt, timedelta

    agent = request.args.get('agent', '').strip()
    date_debut = request.args.get('date_debut', '')
    date_fin = request.args.get('date_fin', '')

    auj = dt.now()
    if not date_debut:
        date_debut = auj.replace(day=1).strftime('%Y-%m-%d')
    if not date_fin:
        fin_mois = (auj.replace(day=1) + timedelta(days=32)).replace(day=1) - timedelta(days=1)
        date_fin = fin_mois.strftime('%Y-%m-%d')

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT DISTINCT agent_nom FROM journal_connexions")
    agents = [row[0] for row in c.fetchall()]

    d1 = dt.strptime(date_debut, '%Y-%m-%d')
    d2 = dt.strptime(date_fin, '%Y-%m-%d')
    jours = []
    d = d1
    while d <= d2:
        jours.append(d.strftime('%Y-%m-%d'))
        d += timedelta(days=1)

    agents_a_exporter = [agent] if agent else agents

    donnees = []
    for ag in agents_a_exporter:
        for jour in jours:
            c.execute("""
                SELECT MIN(date_connexion), MAX(date_connexion)
                FROM journal_connexions
                WHERE agent_nom=? AND date_connexion>=? AND date_connexion<=? AND (type_event='connexion' OR type_event='deconnexion')
            """, (ag, jour+" 00:00:00", jour+" 23:59:59"))
            entree, sortie = c.fetchone()
            heure_entree = entree[11:19] if entree else ''
            heure_sortie = sortie[11:19] if sortie else ''
            if entree and sortie:
                dt1 = dt.strptime(entree, "%Y-%m-%d %H:%M:%S")
                dt2 = dt.strptime(sortie, "%Y-%m-%d %H:%M:%S")
                duree = dt2 - dt1 if dt2 > dt1 else timedelta()
                h = int(duree.total_seconds() // 3600)
                m = int((duree.total_seconds() % 3600) // 60)
                duree_txt = f"{h:02d}:{m:02d}"
            else:
                duree_txt = ''
            donnees.append([ag, jour, heure_entree, heure_sortie, duree_txt])

    df = pd.DataFrame(donnees, columns=["Agent", "Date", "Entrée", "Sortie", "Durée"])
    file_path = "presence_lignes.xlsx"
    df.to_excel(file_path, index=False)
    return send_file(file_path, as_attachment=True)

@app.route('/live_agents')
def live_agents():
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        flash("Accès réservé à l'administration/supervision.", "danger")
        return redirect(url_for('dashboard'))

    now = datetime.now()
    date_auj = now.strftime("%Y-%m-%d")

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT NOM FROM agents")
    agents = [row[0] for row in c.fetchall()]
    live_data = []

    for ag in agents:
        c.execute("""
            SELECT type_event, date_connexion FROM journal_connexions
            WHERE agent_nom=? AND date_connexion>=? ORDER BY date_connexion DESC LIMIT 1
        """, (ag, date_auj + " 00:00:00"))
        last = c.fetchone()
        statut = "Déconnecté"
        heure_statut = ""
        if last:
            type_event, date_evt = last
            heure_statut = date_evt[11:16]
            if type_event == 'connexion':
                statut = "Connecté"
            elif type_event == 'pause':
                statut = "En pause"
            elif type_event == 'saisie':
                statut = "Saisie en cours"
            elif type_event == 'deconnexion':
                statut = "Déconnecté"

        c.execute("""
            SELECT MIN(date_connexion) FROM journal_connexions
            WHERE agent_nom=? AND date_connexion>=? AND type_event='connexion'
        """, (ag, date_auj + " 00:00:00"))
        entree = c.fetchone()[0]
        heure_connexion = entree[11:16] if entree else ""

        c.execute("SELECT COUNT(*) FROM clients WHERE AGENT=? AND DATE_SIGNATURE=?", (ag, date_auj))
        nb_clients = c.fetchone()[0]

        live_data.append({
            "agent": ag,
            "statut": statut,
            "heure_statut": heure_statut,
            "heure_connexion": heure_connexion,
            "nb_clients": nb_clients
        })

    conn.close()
    return render_template('live_agents.html', live_data=live_data, date_auj=date_auj)

@app.route('/api/live_agents')
def api_live_agents():
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        return {"error": "forbidden"}, 403

    now = datetime.now()
    date_auj = now.strftime("%Y-%m-%d")
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT NOM FROM agents")
    agents = [row[0] for row in c.fetchall()]
    live_data = []
    for ag in agents:
        c.execute("""
            SELECT type_event, date_connexion FROM journal_connexions
            WHERE agent_nom=? AND date_connexion>=? ORDER BY date_connexion DESC LIMIT 1
        """, (ag, date_auj + " 00:00:00"))
        last = c.fetchone()
        statut = "Déconnecté"
        heure_statut = ""
        if last:
            type_event, date_evt = last
            heure_statut = date_evt[11:16]
            if type_event == 'connexion':
                statut = "Connecté"
            elif type_event == 'pause':
                statut = "En pause"
            elif type_event == 'saisie':
                statut = "Saisie en cours"
            elif type_event == 'deconnexion':
                statut = "Déconnecté"

        c.execute("""
            SELECT MIN(date_connexion) FROM journal_connexions
            WHERE agent_nom=? AND date_connexion>=? AND type_event='connexion'
        """, (ag, date_auj + " 00:00:00"))
        entree = c.fetchone()[0]
        heure_connexion = entree[11:16] if entree else ""

        c.execute("SELECT COUNT(*) FROM clients WHERE AGENT=? AND DATE_SIGNATURE=?", (ag, date_auj))
        nb_clients = c.fetchone()[0]

        live_data.append({
            "agent": ag,
            "statut": statut,
            "heure_statut": heure_statut,
            "heure_connexion": heure_connexion,
            "nb_clients": nb_clients
        })

    conn.close()
    return {"live_data": live_data}

@app.route('/classement_agents')
def classement_agents():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("""
        SELECT AGENT, COUNT(*) as nb_valides
        FROM clients
        WHERE STATUT = 'valide'
          AND strftime('%Y-%m', DATE_SIGNATURE) = strftime('%Y-%m', 'now')
        GROUP BY AGENT
        ORDER BY nb_valides DESC
        LIMIT 5
    """)
    classement = c.fetchall()
    conn.close()
    total_general = sum([row[1] for row in classement])
    return render_template('classement_agents.html', classement=classement, total_general=total_general)

def notifier_nouveau_client(nom_client):
    socketio.emit('nouveau_client', {'message': f"Nouveau client : {nom_client}"})

@app.route('/overview')
def overview():
    date_debut = request.args.get('date_debut', '').strip()
    date_fin = request.args.get('date_fin', '').strip()
    agent_sfr = request.args.get('agent_sfr', '').strip()
    agent_valandre = request.args.get('agent_valandre', '').strip()

    auj = datetime.now().strftime('%Y-%m-%d')
    if not date_debut and not date_fin:
        date_debut = date_fin = auj
    elif not date_debut:
        date_debut = date_fin
    elif not date_fin:
        date_fin = date_debut

    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()

    c.execute("SELECT DISTINCT AGENT FROM clients WHERE campagne_id=1")
    agents_sfr = sorted({row[0] for row in c.fetchall() if row[0]})
    c.execute("SELECT DISTINCT AGENT FROM clients WHERE campagne_id=2")
    agents_valandre = sorted({row[0] for row in c.fetchall() if row[0]})

    campagne_id_sfr = 1
    params = [campagne_id_sfr, date_debut, date_fin]
    agent_filter = " AND AGENT=?" if agent_sfr else ""
    if agent_sfr: params.append(agent_sfr)
    c.execute(f"""
        SELECT SUM(CASE WHEN STATUT='valide' THEN 1 ELSE 0 END),
               SUM(CASE WHEN STATUT='non valide' THEN 1 ELSE 0 END)
        FROM clients WHERE campagne_id=? AND DATE_SIGNATURE>=? AND DATE_SIGNATURE<=?{agent_filter}
    """, params)
    sfr_valide, sfr_non_valide = c.fetchone() or (0,0)

    c.execute(f"""
        SELECT strftime('%H', DATE_MODIF), COUNT(*)
        FROM clients
        WHERE campagne_id=? AND DATE_SIGNATURE>=? AND DATE_SIGNATURE<=?{agent_filter}
        GROUP BY strftime('%H', DATE_MODIF)
        ORDER BY strftime('%H', DATE_MODIF)
    """, params)
    sfr_par_heure = c.fetchall()

    c.execute(f"""
        SELECT AGENT, strftime('%H', DATE_MODIF), COUNT(*)
        FROM clients
        WHERE campagne_id=? AND DATE_SIGNATURE>=? AND DATE_SIGNATURE<=?{agent_filter}
        GROUP BY AGENT, strftime('%H', DATE_MODIF)
        ORDER BY AGENT, strftime('%H', DATE_MODIF)
    """, params)
    sfr_agent_par_heure = c.fetchall()

    campagne_id_valandre = 2
    params2 = [campagne_id_valandre, date_debut, date_fin]
    agent_filter2 = " AND AGENT=?" if agent_valandre else ""
    if agent_valandre: params2.append(agent_valandre)
    c.execute(f"""
        SELECT
            SUM(CASE WHEN STRATO_STATUT='VALIDÉ' OR LSR_STATUT='VALIDÉ' OR PRESSE_STATUT='VALIDÉ'
                    OR ENI_STATUT='VALIDÉ' OR SERENITY_STATUT='VALIDÉ' OR PROTEC_ALLIANCE_STATUT='VALIDÉ'
                    OR WEKIWI_STATUT='VALIDÉ' THEN 1 ELSE 0 END),
            SUM(CASE WHEN (STRATO_STATUT!='VALIDÉ' AND LSR_STATUT!='VALIDÉ' AND PRESSE_STATUT!='VALIDÉ'
                           AND ENI_STATUT!='VALIDÉ' AND SERENITY_STATUT!='VALIDÉ' AND PROTEC_ALLIANCE_STATUT!='VALIDÉ'
                           AND WEKIWI_STATUT!='VALIDÉ') THEN 1 ELSE 0 END)
        FROM clients WHERE campagne_id=? AND DATE_SIGNATURE>=? AND DATE_SIGNATURE<=?{agent_filter2}
    """, params2)
    valandre_valide, valandre_non_valide = c.fetchone() or (0,0)

    c.execute(f"""
        SELECT strftime('%H', DATE_MODIF), COUNT(*)
        FROM clients
        WHERE campagne_id=? AND DATE_SIGNATURE>=? AND DATE_SIGNATURE<=?{agent_filter2}
        GROUP BY strftime('%H', DATE_MODIF)
        ORDER BY strftime('%H', DATE_MODIF)
    """, params2)
    valandre_par_heure = c.fetchall()

    c.execute(f"""
        SELECT AGENT, strftime('%H', DATE_MODIF), COUNT(*)
        FROM clients
        WHERE campagne_id=? AND DATE_SIGNATURE>=? AND DATE_SIGNATURE<=?{agent_filter2}
        GROUP BY AGENT, strftime('%H', DATE_MODIF)
        ORDER BY AGENT, strftime('%H', DATE_MODIF)
    """, params2)
    valandre_agent_par_heure = c.fetchall()

    conn.close()

    return render_template(
        'overview.html',
        sfr_valide=sfr_valide or 0,
        sfr_non_valide=sfr_non_valide or 0,
        sfr_par_heure=sfr_par_heure,
        sfr_agent_par_heure=sfr_agent_par_heure,
        valandre_valide=valandre_valide or 0,
        valandre_non_valide=valandre_non_valide or 0,
        valandre_par_heure=valandre_par_heure,
        valandre_agent_par_heure=valandre_agent_par_heure,
        agents_sfr=agents_sfr,
        agents_valandre=agents_valandre,
        agent_sfr=agent_sfr,
        agent_valandre=agent_valandre,
        date_debut=date_debut,
        date_fin=date_fin,
        auj=auj
    )

def file_size_okay(file):
    file.seek(0, os.SEEK_END)
    size = file.tell()
    file.seek(0)
    return size <= 5 * 1024 * 1024

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in {'jpg', 'jpeg', 'png'}

# ──────────────────────────────────────────────────────────────────────────────
# Chat SocketIO
# ──────────────────────────────────────────────────────────────────────────────
@socketio.on('chat_message')
def handle_chat_message(data):
    if 'user' in data and 'message' in data:
        conn = sqlite3.connect(DB_NAME)
        c = conn.cursor()
        c.execute("INSERT INTO chat_messages (user, message) VALUES (?, ?)", (data['user'], data['message']))
        conn.commit()
        conn.close()
        emit('chat_message', data, broadcast=True)

@socketio.on('chat_history_request')
def handle_chat_history_request():
    conn = sqlite3.connect(DB_NAME)
    c = conn.cursor()
    c.execute("SELECT user, message, timestamp FROM chat_messages ORDER BY id DESC LIMIT 50")
    rows = c.fetchall()
    conn.close()
    messages = []
    for row in reversed(rows):
        messages.append({'user': row[0], 'message': row[1], 'timestamp': row[2]})
    emit('chat_history', messages)

# ──────────────────────────────────────────────────────────────────────────────
# Export générique
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/export_clients')
def export_clients():
    agent = request.args.get('agent')
    statut = request.args.get('statut')
    date_debut = request.args.get('date_debut')
    date_fin = request.args.get('date_fin')

    sql = "SELECT * FROM clients"
    filters = []
    params = []

    if agent:
        filters.append("AGENT = ?"); params.append(agent)
    if statut:
        filters.append("STATUT = ?"); params.append(statut)
    if date_debut and date_fin:
        filters.append("DATE_SIGNATURE BETWEEN ? AND ?"); params.extend([date_debut, date_fin])
    elif date_debut:
        filters.append("DATE_SIGNATURE >= ?"); params.append(date_debut)
    elif date_fin:
        filters.append("DATE_SIGNATURE <= ?"); params.append(date_fin)

    if filters:
        sql += " WHERE " + " AND ".join(filters)
    sql += " ORDER BY DATE_SIGNATURE DESC"

    conn = sqlite3.connect(DB_NAME)
    df = pd.read_sql_query(sql, conn, params=params)
    conn.close()

    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name="Clients")
    output.seek(0)
    return send_file(output, download_name="export_clients.xlsx", as_attachment=True)

# ──────────────────────────────────────────────────────────────────────────────
# Debug / Télé
# ──────────────────────────────────────────────────────────────────────────────
# Debug / Téléchargement Aircall
# ──────────────────────────────────────────────────────────────────────────────

@app.route('/telecharger_aircall_numero/<phone_number>', methods=['POST'])
def telecharger_aircall_numero(phone_number):
    if 'agent_nom' not in session:
        flash("Session expirée.", "warning")
        return redirect(url_for('login'))

    clean = _normalize_phone(phone_number)

    try:
        # 1) Toujours re-demander l'URL la plus récente à Aircall (pas de cache)
        recording_url = find_recording_for_phone_number(clean)
        if not recording_url:
            # Si pas d'enregistrement, on essaie au moins de rafraîchir le CALL_ID
            update_call_id_in_db(clean)
            flash("Aucun enregistrement trouvé pour ce numéro.", "warning")
            return redirect(url_for('dashboard'))

        # 2) Télécharger le flux à chaud
        r = requests.get(recording_url, timeout=60)
        if r.status_code != 200:
            flash(f"Téléchargement impossible (HTTP {r.status_code}).", "danger")
            return redirect(url_for('dashboard'))

        # 3) Nom de fichier unique à chaque requête
        unique_id = uuid.uuid4().hex[:8]
        safe_number = re.sub(r"[^\d+]", "_", clean)
        filename = f"aircall_{safe_number}_{unique_id}.mp3"

        # 4) Mettre à jour CALL_ID en base (si on l'a)
        update_call_id_in_db(clean)

        buf = BytesIO(r.content)
        buf.seek(0)

        resp = send_file(buf, as_attachment=True, download_name=filename, mimetype='audio/mpeg')
        # Désactive tout cache navigateur / proxy
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp

    except Exception as e:
        print(f"[telecharger_aircall_numero] Erreur: {e}")
        flash("Erreur lors du téléchargement de l'enregistrement.", "danger")
        return redirect(url_for('dashboard'))


@app.route('/debug_aircall/<phone_number>')
def debug_aircall(phone_number):
    """Debug: affiche les infos trouvées pour un numéro (restreint admin/superviseur)."""
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        return jsonify({"error": "Accès interdit"}), 403

    info = {
        "phone_number": str(phone_number),
        "timestamp": datetime.now().isoformat()
    }
    try:
        recording_url = find_recording_for_phone_number(phone_number)
        info["recording_url"] = recording_url
        if recording_url:
            try:
                head = requests.head(recording_url, timeout=10)
                info["url_status"] = head.status_code
                info["content_type"] = head.headers.get('Content-Type')
                info["content_length"] = head.headers.get('Content-Length')
            except Exception as e:
                info["url_head_error"] = str(e)
    except Exception as e:
        info["error"] = str(e)
        return jsonify(info), 500

    return jsonify(info)


@app.route('/telecharger_aircall_test/<phone_number>', methods=['POST'])
def telecharger_aircall_test(phone_number):
    """Version de test (sans cache) pour télécharger l'enregistrement d'un numéro."""
    if 'agent_nom' not in session:
        return redirect(url_for('login'))

    try:
        recording_url = find_recording_for_phone_number(phone_number)
        if not recording_url:
            flash("Aucun enregistrement trouvé pour ce numéro de téléphone.", "warning")
            return redirect(url_for('dashboard'))

        resp = requests.get(recording_url, timeout=60)
        if resp.status_code == 200:
            safe_number = re.sub(r"[^\d+]", "_", str(phone_number))
            unique_id = str(uuid.uuid4())[:8]
            filename = f"test_{safe_number}_{unique_id}.mp3"

            file_stream = BytesIO(resp.content)
            file_stream.seek(0)
            return send_file(
                file_stream,
                as_attachment=True,
                download_name=filename,
                mimetype='audio/mpeg'
            )
        else:
            flash(f"Erreur HTTP {resp.status_code}", "danger")
    except Exception as e:
        print(f"[telecharger_aircall_test] Erreur: {e}")
        flash(f"Erreur test : {e}", "danger")

    return redirect(url_for('dashboard'))


@app.route('/clear_aircall_cache', methods=['POST'])
def clear_aircall_cache():
    """(Placeholder) Vide les caches potentiels côté serveur."""
    if session.get('agent_role', '') not in ['admin', 'superviseur']:
        return jsonify({"error": "Accès interdit"}), 403

    import gc
    gc.collect()
    flash("Cache vidé (si applicable).", "success")
    return redirect(url_for('dashboard'))

# ──────────────────────────────────────────────────────────────────────────────
# Route API: forcer la résolution du CALL_ID depuis Aircall par numéro
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/resolve_call_id/<phone_number>', methods=['POST', 'GET'])
def resolve_call_id(phone_number):
    if 'agent_nom' not in session:
        return jsonify({"error": "Unauthorized"}), 401

    clean = _normalize_phone(phone_number)
    call_id = update_call_id_in_db(clean)

    return jsonify({
        "phone_number": clean,
        "call_id": call_id,
        "updated": bool(call_id)
    })

# ──────────────────────────────────────────────────────────────────────────────
# (Optionnel) Backfill admin: remplir les CALL_ID manquants pour tous les clients
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/admin/backfill_call_ids', methods=['POST'])
def backfill_call_ids():
    if session.get('agent_role') not in ['admin', 'superviseur']:
        return jsonify({"error": "forbidden"}), 403

    conn = sqlite3.connect(DB_NAME); c = conn.cursor()
    c.execute("""
        SELECT DISTINCT TELEPHONE
          FROM clients
         WHERE TELEPHONE IS NOT NULL AND TELEPHONE <> ''
           AND (CALL_ID IS NULL OR CALL_ID = '')
    """)
    phones = [row[0] for row in c.fetchall()]
    conn.close()

    updated = []
    for p in phones:
        cid = update_call_id_in_db(p)
        if cid:
            updated.append({"phone": _normalize_phone(p), "call_id": cid})

    return jsonify({"updated_count": len(updated), "details": updated})

# ──────────────────────────────────────────────────────────────────────────────
# Télécharger par CALL_ID (bypass complet du cache, nom unique à chaque fois)
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/telecharger_aircall_call/<call_id>', methods=['POST', 'GET'])
def telecharger_aircall_call(call_id):
    if 'agent_nom' not in session:
        flash("Session expirée.", "warning")
        return redirect(url_for('login'))

    try:
        # 1) On récupère l'URL d'enregistrement depuis l'appel Aircall
        #    → évite toute ambiguïté si plusieurs contacts partagent un n°
        r_call = requests.get(
            f"https://api.aircall.io/v1/calls/{call_id}",
            auth=HTTPBasicAuth(API_ID, API_TOKEN),
            timeout=20
        )
        r_call.raise_for_status()
        call_data = r_call.json().get('call') or {}
        rec_url = call_data.get('recording')
        if not rec_url:
            flash("Aucun enregistrement pour cet appel.", "warning")
            return redirect(url_for('dashboard'))

        # 2) Téléchargement 'à chaud' de l'asset (pas de cache!)
        r_file = requests.get(rec_url, timeout=60)
        if r_file.status_code != 200:
            flash(f"Impossible de télécharger (HTTP {r_file.status_code}).", "danger")
            return redirect(url_for('dashboard'))

        unique_id = uuid.uuid4().hex[:8]
        filename = f"aircall_call_{call_id}_{unique_id}.mp3"

        buf = BytesIO(r_file.content); buf.seek(0)
        resp = send_file(buf, as_attachment=True, download_name=filename, mimetype='audio/mpeg')
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        return resp

    except Exception as e:
        print(f"[telecharger_aircall_call] Erreur: {e}")
        flash("Erreur lors du téléchargement de l'enregistrement.", "danger")
        return redirect(url_for('dashboard'))


# ──────────────────────────────────────────────────────────────────────────────
# Rafraîchir/poser un CALL_ID pour 1 client (depuis son id en base)
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/resolve_call_id_for_client/<int:client_id>', methods=['POST'])
def resolve_call_id_for_client(client_id):
    if 'agent_nom' not in session:
        return jsonify({"error": "Unauthorized"}), 401
    try:
        conn = sqlite3.connect(DB_NAME); c = conn.cursor()
        c.execute("SELECT TELEPHONE FROM clients WHERE id=?", (client_id,))
        row = c.fetchone()
        conn.close()
        if not row or not row[0]:
            return jsonify({"client_id": client_id, "updated": False, "reason": "no phone"}), 200

        phone = _normalize_phone(row[0])
        call_id = update_call_id_in_db(phone)
        return jsonify({"client_id": client_id, "phone": phone, "call_id": call_id, "updated": bool(call_id)}), 200
    except Exception as e:
        return jsonify({"client_id": client_id, "error": str(e)}), 500


# ──────────────────────────────────────────────────────────────────────────────
# Lecture/stream inline (pour tester sans téléchargement, et éviter le cache)
# ──────────────────────────────────────────────────────────────────────────────
@app.route('/play_aircall/<phone_number>')
def play_aircall(phone_number):
    if 'agent_nom' not in session:
        return redirect(url_for('login'))
    try:
        rec_url = find_recording_for_phone_number(_normalize_phone(phone_number))
        if not rec_url:
            flash("Aucun enregistrement trouvé pour ce numéro.", "warning")
            return redirect(url_for('dashboard'))

        r = requests.get(rec_url, timeout=60)
        if r.status_code != 200:
            flash(f"Lecture impossible (HTTP {r.status_code}).", "danger")
            return redirect(url_for('dashboard'))

        buf = BytesIO(r.content); buf.seek(0)
        resp = send_file(buf, as_attachment=False, download_name="aircall_preview.mp3", mimetype='audio/mpeg')
        resp.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, max-age=0'
        resp.headers['Pragma'] = 'no-cache'
        resp.headers['Expires'] = '0'
        # Empêche le navigateur de "mémoriser" l'URL par défaut
        resp.headers['Content-Disposition'] = 'inline; filename="aircall_preview.mp3"'
        return resp
    except Exception as e:
        print(f"[play_aircall] Erreur: {e}")
        flash("Erreur pendant la lecture de l'enregistrement.", "danger")
        return redirect(url_for('dashboard'))



@app.errorhandler(429)
def ratelimit_handler(e):
    # Petit message, et on renvoie vers la page de login
    flash("Trop de tentatives de connexion. Réessayez dans 1 minute.", "danger")
    return redirect(url_for('login'))


# ──────────────────────────────────────────────────────────────────────────────
# Entrée principale
# ──────────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    socketio.run(app, debug=True)
